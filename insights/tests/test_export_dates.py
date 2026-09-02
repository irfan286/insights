"""A date leaves the export as a date, not as the number a sheet stores it as.

A spreadsheet holds 2024-01-01 as the serial 45292 and decides what to show
from the cell's number format. A cell written without one shows the serial, so
the reader has to select the column and format it by hand after every
download. The writer is the last place that still knows which columns held
dates.
"""

import base64
import datetime
from io import BytesIO
from unittest.mock import MagicMock, patch

import frappe
import openpyxl
import pandas as pd
from frappe.tests import UnitTestCase

from insights.tests.base import InsightsIntegrationTestCase
from insights.tests.factories import DT, create_test_query, create_test_workbook, delete_users
from insights.tests.permissions_utils import ADMIN, create_test_users
from insights.utils import date_number_formats


def formats(date_format="dd-mm-yyyy", **columns):
    """The formats an export of these columns would carry, on a site set to `date_format`."""
    with patch("insights.utils.frappe.db.get_single_value", return_value=date_format):
        return date_number_formats(pd.DataFrame(columns))


class TestDateColumnsCarryANumberFormat(UnitTestCase):
    def test_a_date_column_is_formatted_as_a_date(self):
        """ibis hands a date column back as `object`, holding date instances."""
        self.assertEqual(formats(day=[datetime.date(2024, 1, 1)]), ["dd-mm-yyyy"])

    def test_a_timestamp_column_keeps_its_time(self):
        self.assertEqual(
            formats(at=pd.to_datetime(["2024-01-01 10:30:00"])),
            ["dd-mm-yyyy hh:mm:ss"],
        )

    def test_a_timestamp_column_of_midnights_reads_as_a_date(self):
        """A source storing a date as a timestamp shouldn't add 00:00:00 to every row."""
        self.assertEqual(formats(at=pd.to_datetime(["2024-01-01", "2024-02-03"])), ["dd-mm-yyyy"])

    def test_a_time_column_is_formatted_as_a_time(self):
        self.assertEqual(formats(clock=[datetime.time(9, 30)]), ["hh:mm:ss"])

    def test_a_column_is_read_past_its_nulls(self):
        self.assertEqual(formats(day=[None, datetime.date(2024, 1, 1)]), ["dd-mm-yyyy"])

    def test_the_site_date_format_is_the_one_written(self):
        self.assertEqual(formats("yyyy-mm-dd", day=[datetime.date(2024, 1, 1)]), ["yyyy-mm-dd"])

    def test_an_unknown_site_date_format_falls_back_to_iso(self):
        """The string is a number format code, so one a sheet can't read is not written."""
        self.assertEqual(formats("dd MMM yy", day=[datetime.date(2024, 1, 1)]), ["yyyy-mm-dd"])


class TestOtherColumnsAreLeftAlone(UnitTestCase):
    def test_a_number_column_carries_no_format(self):
        """A number format on a number would pin how the reader sees it."""
        self.assertEqual(formats(amount=[1.5, 2]), [None])

    def test_a_text_column_carries_no_format(self):
        self.assertEqual(formats(name=["Ada"]), [None])

    def test_an_empty_column_carries_no_format(self):
        self.assertEqual(formats(unknown=[None, None]), [None])

    def test_the_formats_are_read_by_position(self):
        """Two columns of an export may share a label; cells are addressed by index."""
        frame = pd.DataFrame([[datetime.date(2024, 1, 1), "Ada"]], columns=["day", "day"])
        with patch("insights.utils.frappe.db.get_single_value", return_value="dd-mm-yyyy"):
            self.assertEqual(date_number_formats(frame), ["dd-mm-yyyy", None])


class TestTheExcelDownloadFormatsDates(InsightsIntegrationTestCase):
    """The formats are worth nothing if the writer skips them."""

    @classmethod
    def before_class(cls):
        create_test_users()
        cls.workbook = create_test_workbook(ADMIN, title="Export Dates Workbook").name
        cls.query = create_test_query(ADMIN, cls.workbook, title="Export Dates Query").name

    @classmethod
    def after_class(cls):
        frappe.delete_doc(DT.WORKBOOK, cls.workbook, force=True, ignore_permissions=True)
        delete_users(ADMIN)

    def sheet(self):
        doc = frappe.get_doc(DT.QUERY, self.query)
        rows = pd.DataFrame(
            {
                "day": [datetime.date(2024, 1, 1)],
                "at": pd.to_datetime(["2024-01-01 10:30:00"]),
                "amount": [1.5],
            }
        )
        built = MagicMock()
        built.columns = []
        built.limit.return_value = built
        with (
            patch.object(type(doc), "build", return_value=built),
            patch(
                "insights.insights.doctype.insights_query_v3.insights_query_v3.execute_ibis_query",
                return_value=(rows, 0),
            ),
        ):
            workbook = base64.b64decode(doc.download_results(format="excel"))
        return openpyxl.load_workbook(BytesIO(workbook)).active

    def test_a_date_cell_is_written_as_a_date(self):
        cell = self.sheet().cell(row=2, column=1)
        self.assertIsInstance(cell.value, datetime.datetime)
        self.assertNotEqual(cell.number_format, "General")

    def test_a_timestamp_cell_keeps_its_time(self):
        cell = self.sheet().cell(row=2, column=2)
        self.assertIn("hh:mm", cell.number_format)
        self.assertEqual(cell.value.hour, 10)

    def test_a_number_cell_is_untouched(self):
        self.assertEqual(self.sheet().cell(row=2, column=3).number_format, "General")

    def test_a_date_column_is_wide_enough_to_show_the_date(self):
        """Too narrow, and the sheet shows ### where the date was."""
        self.assertGreater(self.sheet().column_dimensions["A"].width, 10)
